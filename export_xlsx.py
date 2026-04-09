from openpyxl import Workbook
from typing import List, Dict
from pathlib import Path
import json
import openpyxl
from openpyxl.utils import get_column_letter


def build_workbook_for_contracts(contracts: List[Dict]) -> Workbook:
    """Create an openpyxl Workbook from a list of contract dicts.

    Columns include a broad selection of contract fields.
    """
    wb = openpyxl.load_workbook('./data/sample.xlsm', keep_vba=True)
    ws = wb.active
    ws.title = "合同信息"

    # 更宽泛的表头集合，不被源码字段限制
    headers = [
        "合同编号",
        "合同名称",
        "合同类型",
        "合同分类",
        "甲方",
        "乙方",
        "甲方代表",
        "甲方电话",
        "乙方代表",
        "合同金额",
        "签订日期",
        "合同开始日期",
        "合同结束日期",
        "部门",
        "项目名称",
        "操作人",
        "创建公司",
        "创建人",
        "创建时间",
        "更新时间",
        "省份",
        "区域",
        "购买数量",
        "标准数量",
        "利润率",
        "业务类型"
    ]
    ws.append(headers)

    for c in contracts:
        # 使用多个候选字段以提高兼容性
        row = [
            c.get("contractCode") or c.get("contractNo") or "",
            c.get("contractName", ""),
            c.get("contractTypeName") or c.get("contractType", ""),
            c.get("contractClassifyName") or c.get("contractClassify", ""),
            c.get("partyA") or c.get("purchaserName") or c.get("partyAName") or "",
            c.get("partyB") or c.get("supplierName") or c.get("partyBName") or "",
            c.get("partyARepresentative", ""),
            c.get("partyAPhoneNumber", ""),
            c.get("partyBRepresentative", ""),
            c.get("contractAmount", ""),
            c.get("signDate", ""),
            c.get("contractStartDate", ""),
            c.get("contractEndDate") or c.get("contractEndDateStr", ""),
            c.get("deptName") or c.get("deptCode") or "",
            c.get("projectName", ""),
            c.get("operatorName") or c.get("employeeName") or "",
            c.get("createCompanyName") or c.get("createCompanyFullName") or "",
            c.get("createUserName", ""),
            c.get("createTime", ""),
            c.get("updateTime", ""),
            c.get("provinceName", ""),
            c.get("areaName", ""),
            c.get("buyCount", ""),
            c.get("stdCount", ""),
            c.get("profitRate", ""),
            c.get("gdContractType") or c.get("businessType") or ""
        ]
        ws.append(row)

    # 自动调整列宽和行高
    _auto_adjust_dimensions(ws)

    return wb


def _auto_adjust_dimensions(ws) -> None:
    """自动调整工作表的列宽和行高。
    
    列宽基于该列最长内容的长度计算。
    行高基于内容中的换行符数量自适应调整。
    """
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                # 计算单元格内容的长度（考虑汉字占用较宽）
                cell_value = str(cell.value) if cell.value is not None else ""
                # 统计字符数，汉字按 2 个字符宽度计算
                display_length = sum(4 if ord(char) > 127 else 2 for char in cell_value)
                if display_length > max_length:
                    max_length = display_length
            except:
                pass
        
        # 设置列宽（宽度 = 字符长度 / 2 + 2 作为外边距）
        adjusted_width = (max_length / 2) + 2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 最大宽度不超过 50
    
    # 调整行高
    for row_num, row in enumerate(ws.iter_rows(), 1):
        max_height = 15  # 最小行高
        
        for cell in row:
            if cell.value:
                # 统计换行符数量
                line_count = str(cell.value).count('\n') + 1
                # 每行增加 15 点高度
                cell_height = line_count * 15
                if cell_height > max_height:
                    max_height = cell_height
        
        ws.row_dimensions[row_num].height = max_height


def export_contracts_to_xlsm(contracts: List[Dict], output_filename: str = "contracts_export.xlsm") -> str:
    """Export a list of contracts to an xlsm file with a default filename.

    Returns the path to the saved file.
    """
    wb = build_workbook_for_contracts(contracts)
    wb.save(output_filename)
    return str(Path(output_filename).resolve())


if __name__ == '__main__':
    # 当作为脚本运行时，尝试读取 data/selected_contract.json 并导出为 xlsm
    data_path = Path(__file__).parent / 'data' / 'selected_contract.json'
    if data_path.exists():
        with open(data_path, 'r', encoding='utf-8') as f:
            contract = json.load(f)
        out = export_contracts_to_xlsm([contract])
        print(f"Exported contracts to: {out}")
    else:
        print("selected_contract.json not found in data/")